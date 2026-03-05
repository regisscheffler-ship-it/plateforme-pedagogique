"""
Script d'import des référentiels depuis le fichier Excel.
VERSION CORRIGÉE — les CompetenceProfessionnelle sont créées pour CHAQUE bloc
où la compétence parente existe (ex: C2.1 dans E1, E2 et E3).

Placer ce fichier à la racine du projet Django et lancer :
    python import_referentiels.py
"""

import os
import django
import sys

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'plateforme.settings')
django.setup()

from core.models import (
    Referentiel, BlocCompetence, Competence,
    CompetenceProfessionnelle, SousCompetence,
    CritereEvaluation, IndicateurPerformance, Connaissance
)

try:
    import openpyxl
except ImportError:
    print("❌ openpyxl non installé. Lance : pip install openpyxl")
    sys.exit(1)

EXCEL_FILE = os.path.join(os.path.dirname(__file__), 'Referentiels_Loucheur.xlsx')

if not os.path.exists(EXCEL_FILE):
    print(f"❌ Fichier non trouvé : {EXCEL_FILE}")
    sys.exit(1)

print(f"📂 Lecture du fichier : {EXCEL_FILE}")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)


def get_sheet(name):
    for sheet_name in wb.sheetnames:
        if sheet_name.upper() == name.upper():
            return wb[sheet_name]
    for sheet_name in wb.sheetnames:
        if name.upper() in sheet_name.upper():
            return wb[sheet_name]
    raise ValueError(f"Feuille '{name}' non trouvée. Disponibles : {wb.sheetnames}")


def sheet_to_dicts(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        d = {headers[i]: (str(row[i]).strip() if row[i] is not None else '') for i in range(len(headers))}
        result.append(d)
    return result


# ============================================================
# LECTURE
# ============================================================
print("\n📖 Lecture des feuilles Excel...")

blocs_data         = sheet_to_dicts(get_sheet('1_BLOCS'))
competences_data   = sheet_to_dicts(get_sheet('2_COMPETENCES'))
cp_data            = sheet_to_dicts(get_sheet('3_COMPETENCES_PRO'))
sc_data            = sheet_to_dicts(get_sheet('4_SOUS_COMPETENCES'))
criteres_data      = sheet_to_dicts(get_sheet('5_CRITERES'))
connaissances_data = sheet_to_dicts(get_sheet('6_CONNAISSANCES'))

print(f"  ✅ Blocs           : {len(blocs_data)}")
print(f"  ✅ Compétences     : {len(competences_data)}")
print(f"  ✅ Compétences Pro : {len(cp_data)}")
print(f"  ✅ Sous-compétences: {len(sc_data)}")
print(f"  ✅ Critères        : {len(criteres_data)}")
print(f"  ✅ Connaissances   : {len(connaissances_data)}")


# ============================================================
# ÉTAPE 1 : BLOCS
# ============================================================
print("\n📦 Import des Blocs de compétences...")

blocs_map = {}  # (code, ref_nom) -> BlocCompetence
blocs_created = blocs_existing = 0

for row in blocs_data:
    code    = row.get('code', '').strip()
    nom     = row.get('nom', '').strip()
    ref_nom = row.get('referentiel', '').strip()
    ordre   = int(row.get('ordre', 0) or 0)

    if not code or not nom or not ref_nom:
        continue

    try:
        ref = Referentiel.objects.get(nom=ref_nom)
    except Referentiel.DoesNotExist:
        print(f"  ⚠️  Référentiel '{ref_nom}' non trouvé, ignoré.")
        continue

    bloc, created = BlocCompetence.objects.get_or_create(
        code=code, referentiel=ref,
        defaults={'nom': nom, 'ordre': ordre}
    )
    if not created and (bloc.nom != nom or bloc.ordre != ordre):
        bloc.nom = nom
        bloc.ordre = ordre
        bloc.save()

    blocs_map[(code, ref_nom)] = bloc
    if created:
        blocs_created += 1
    else:
        blocs_existing += 1

print(f"  ✅ {blocs_created} créés, {blocs_existing} déjà existants")


# ============================================================
# ÉTAPE 2 : COMPÉTENCES
# ============================================================
print("\n🎯 Import des Compétences...")

# ✅ CLÉ CORRIGÉE : (code, bloc_code, ref_nom) pour distinguer C2/E1 de C2/E2 de C2/E3
competences_map = {}  # (code, bloc_code, ref_nom) -> Competence
comp_created = comp_existing = 0

for row in competences_data:
    code      = row.get('code', '').strip()
    nom       = row.get('nom', '').strip()
    bloc_code = row.get('bloc_code', '').strip()
    ref_nom   = row.get('referentiel', '').strip()
    ordre     = int(row.get('ordre', 0) or 0)

    if not code or not nom or not bloc_code or not ref_nom:
        continue

    bloc = blocs_map.get((bloc_code, ref_nom))
    if not bloc:
        print(f"  ⚠️  Bloc '{bloc_code}' non trouvé pour '{ref_nom}', compétence '{code}' ignorée.")
        continue

    comp, created = Competence.objects.get_or_create(
        code=code, bloc=bloc,
        defaults={'nom': nom, 'ordre': ordre}
    )
    if not created and comp.nom != nom:
        comp.nom = nom
        comp.save()

    competences_map[(code, bloc_code, ref_nom)] = comp
    if created:
        comp_created += 1
    else:
        comp_existing += 1

print(f"  ✅ {comp_created} créées, {comp_existing} déjà existantes")


# ============================================================
# ÉTAPE 3 : COMPÉTENCES PROFESSIONNELLES
# ✅ CORRECTION PRINCIPALE : on crée une CP pour CHAQUE bloc
#    où la compétence parente existe, pas juste le premier trouvé.
# ============================================================
print("\n🏆 Import des Compétences Professionnelles...")

# cp_map : (code, bloc_code, ref_nom) -> CompetenceProfessionnelle
# Ainsi C2.1 dans E1, E2 et E3 sont bien trois entrées distinctes
cp_map = {}
cp_created = cp_existing = 0

for row in cp_data:
    code      = row.get('code', '').strip()
    nom       = row.get('nom', '').strip()
    comp_code = row.get('competences_code', '').strip()
    ref_nom   = row.get('referentiel', '').strip()
    ordre     = int(row.get('ordre', 0) or 0)

    if not code or not nom or not comp_code or not ref_nom:
        continue

    # ✅ Cherche TOUTES les compétences parentes correspondantes
    #    (C2 peut exister dans E1, E2, E3 → on crée la CP dans chacun)
    comp_matches = [
        (k, comp) for k, comp in competences_map.items()
        if k[0] == comp_code and k[2] == ref_nom
    ]

    if not comp_matches:
        print(f"  ⚠️  Compétence '{comp_code}' non trouvée pour '{ref_nom}', CP '{code}' ignorée.")
        continue

    for (comp_code_key, bloc_code_key, ref_key), comp_obj in comp_matches:
        cp, created = CompetenceProfessionnelle.objects.get_or_create(
            code=code,
            competence=comp_obj,
            defaults={'nom': nom, 'ordre': ordre}
        )
        if not created and cp.nom != nom:
            cp.nom = nom
            cp.save()

        # ✅ Clé avec le bloc pour distinguer C2.1/E1 de C2.1/E2 de C2.1/E3
        cp_map[(code, bloc_code_key, ref_nom)] = cp

        if created:
            cp_created += 1
            print(f"  ➕ CP '{code}' créée dans bloc '{bloc_code_key}' ({ref_nom})")
        else:
            cp_existing += 1

print(f"  ✅ {cp_created} créées, {cp_existing} déjà existantes")


# ============================================================
# ÉTAPE 4 : SOUS-COMPÉTENCES
# ✅ CORRIGÉ : cherche dans cp_map avec le bon bloc
# ============================================================
print("\n🔧 Import des Sous-Compétences...")

# sc_map : (code, bloc_code, ref_nom) -> SousCompetence
sc_map = {}
sc_created = sc_existing = 0

for row in sc_data:
    code    = row.get('code', '').strip()
    nom     = row.get('nom', '').strip()
    cp_code = row.get('cp_code', '').strip()
    ref_nom = row.get('referentiel', '').strip()
    ordre   = int(row.get('ordre', 0) or 0)

    if not code or not nom or not cp_code or not ref_nom:
        continue

    # ✅ Cherche toutes les CP correspondantes (une par bloc)
    cp_matches = [
        (bloc_code, cp_obj) for (cp_c, bloc_code, ref), cp_obj in cp_map.items()
        if cp_c == cp_code and ref == ref_nom
    ]

    if not cp_matches:
        print(f"  ⚠️  CP '{cp_code}' non trouvée pour '{ref_nom}', SC '{code}' ignorée.")
        continue

    for bloc_code, cp_obj in cp_matches:
        sc, created = SousCompetence.objects.get_or_create(
            code=code,
            competence_pro=cp_obj,
            defaults={'nom': nom, 'ordre': ordre}
        )
        if not created and sc.nom != nom:
            sc.nom = nom
            sc.save()

        sc_map[(code, bloc_code, ref_nom)] = sc

        if created:
            sc_created += 1
        else:
            sc_existing += 1

print(f"  ✅ {sc_created} créées, {sc_existing} déjà existantes")


# ============================================================
# ÉTAPE 5 : CRITÈRES
# ✅ CORRIGÉ : cherche dans sc_map avec le bon bloc
# ============================================================
print("\n📋 Import des Critères d'évaluation...")

# crit_map : (code, bloc_code, ref_nom) -> CritereEvaluation
crit_map = {}
criteres_created = criteres_existing = 0

for row in criteres_data:
    code    = row.get('code', '').strip()
    nom     = row.get('nom', '').strip()
    sc_code = row.get('sc_code', '').strip()
    ref_nom = row.get('referentiel', '').strip()
    ordre   = int(row.get('ordre', 0) or 0)

    if not code or not nom or not sc_code or not ref_nom:
        continue

    # ✅ Cherche toutes les SC correspondantes (une par bloc)
    sc_matches = [
        (bloc_code, sc_obj) for (sc_c, bloc_code, ref), sc_obj in sc_map.items()
        if sc_c == sc_code and ref == ref_nom
    ]

    if not sc_matches:
        print(f"  ⚠️  SC '{sc_code}' non trouvée pour '{ref_nom}', critère '{code}' ignoré.")
        continue

    for bloc_code, sc_obj in sc_matches:
        crit, created = CritereEvaluation.objects.get_or_create(
            code=code,
            sous_competence=sc_obj,
            defaults={'nom': nom, 'ordre': ordre}
        )
        if not created and crit.nom != nom:
            crit.nom = nom
            crit.save()

        crit_map[(code, bloc_code, ref_nom)] = crit

        if created:
            criteres_created += 1
        else:
            criteres_existing += 1

print(f"  ✅ {criteres_created} créés, {criteres_existing} déjà existants")


# ============================================================
# ÉTAPE 6 : INDICATEURS DE PERFORMANCE
# (un indicateur par critère — créé automatiquement s'il n'existe pas)
# ============================================================
print("\n🎯 Création des Indicateurs de performance (1 par critère)...")

ind_created = ind_existing = 0

for (code, bloc_code, ref_nom), crit in crit_map.items():
    ind, created = IndicateurPerformance.objects.get_or_create(
        critere=crit,
        defaults={'nom': crit.nom, 'ordre': 1}
    )
    if created:
        ind_created += 1
    else:
        ind_existing += 1

print(f"  ✅ {ind_created} créés, {ind_existing} déjà existants")


# ============================================================
# ÉTAPE 7 : CONNAISSANCES
# ✅ CORRIGÉ : rattache à TOUTES les CP correspondantes (un par bloc)
# ✅ DÉDUPLICATION : pas de doublon pour la même CP
# ============================================================
print("\n📚 Import des Connaissances...")

conn_created = conn_existing = 0

for row in connaissances_data:
    code    = row.get('code', '').strip()
    nom     = row.get('nom', '').strip()
    cp_code = row.get('cp_code', '').strip()
    ref_nom = row.get('referentiel', '').strip()
    ordre   = int(row.get('ordre', 0) or 0)

    if not code or not nom or not cp_code or not ref_nom:
        continue

    # ✅ Cherche toutes les CP correspondantes (une par bloc)
    cp_matches = [
        cp_obj for (cp_c, bloc_code, ref), cp_obj in cp_map.items()
        if cp_c == cp_code and ref == ref_nom
    ]

    if not cp_matches:
        print(f"  ⚠️  CP '{cp_code}' non trouvée pour '{ref_nom}', connaissance '{code}' ignorée.")
        continue

    for cp_obj in cp_matches:
        conn, created = Connaissance.objects.get_or_create(
            nom=nom,
            competence_pro=cp_obj,
            defaults={'ordre': ordre}
        )
        if created:
            conn_created += 1
        else:
            conn_existing += 1

print(f"  ✅ {conn_created} créées, {conn_existing} déjà existantes")


# ============================================================
# RÉSUMÉ FINAL
# ============================================================
print("\n" + "=" * 60)
print("✅ IMPORT TERMINÉ !")
print("=" * 60)

for ref in Referentiel.objects.all():
    nb_blocs = BlocCompetence.objects.filter(referentiel=ref).count()
    nb_comp  = Competence.objects.filter(bloc__referentiel=ref).count()
    nb_cp    = CompetenceProfessionnelle.objects.filter(competence__bloc__referentiel=ref).count()
    nb_sc    = SousCompetence.objects.filter(competence_pro__competence__bloc__referentiel=ref).count()
    nb_crit  = CritereEvaluation.objects.filter(sous_competence__competence_pro__competence__bloc__referentiel=ref).count()
    nb_conn  = Connaissance.objects.filter(competence_pro__competence__bloc__referentiel=ref).count()
    nb_ind   = IndicateurPerformance.objects.filter(critere__sous_competence__competence_pro__competence__bloc__referentiel=ref).count()
    print(f"\n📚 {ref.nom} (ID={ref.id})")
    print(f"   Blocs: {nb_blocs} | Compétences: {nb_comp} | CP: {nb_cp}")
    print(f"   SC: {nb_sc} | Critères: {nb_crit} | Indicateurs: {nb_ind} | Connaissances: {nb_conn}")

print()
print("🔍 Vérification C2 dans CAP Maçon 2021 :")
from core.models import CompetenceProfessionnelle
for cp in CompetenceProfessionnelle.objects.filter(
    code__startswith='C2',
    competence__bloc__referentiel__nom='CAP Maçon 2021'
).select_related('competence__bloc'):
    print(f"   {cp.code} | {cp.nom[:40]} | bloc: {cp.competence.bloc.code}")
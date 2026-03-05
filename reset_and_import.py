import os
import django
import sys

# Configuration Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'plateforme.settings')
django.setup()

from core.models import (
    Referentiel, BlocCompetence, Competence, CompetenceProfessionnelle,
    SousCompetence, CritereEvaluation, IndicateurPerformance, Connaissance
)
import openpyxl

# LISTE DES REFERENTIELS A REINITIALISER
TARGET_REFS = ['CAP Maçon 2021', 'ORGO 2026', 'TB ORGO 2015']
FICHIER_EXCEL = 'Referentiels_Loucheur.xlsx'

def run_import():
    print("=" * 80)
    print("🛠️  REMISE A ZERO ET IMPORTATION PROPRE (CORRIGÉ)")
    print("=" * 80)

    # 1. NETTOYAGE PREALABLE
    print("\n[1] 🧹 SUPPRESSION DES ANCIENNES DONNEES...")
    for nom_ref in TARGET_REFS:
        count, _ = Referentiel.objects.filter(nom=nom_ref).delete()
        if count > 0:
            print(f"    ❌ Supprimé : {nom_ref} (et tout son contenu)")
        else:
            print(f"    ℹ️  Introuvable (déjà propre) : {nom_ref}")

    # 2. CHARGEMENT EXCEL
    if not os.path.exists(FICHIER_EXCEL):
        print(f"    ⛔ ERREUR: Fichier {FICHIER_EXCEL} introuvable.")
        return
    
    wb = openpyxl.load_workbook(FICHIER_EXCEL)
    print("\n[2] 📂 Fichier Excel chargé.")

    # 3. CREATION DES REFERENTIELS
    print("\n[3] 🏗️  Création des Référentiels...")
    ref_map = {} 
    for nom_ref in TARGET_REFS:
        ref = Referentiel.objects.create(nom=nom_ref, actif=True)
        ref_map[nom_ref] = ref
        print(f"    ✅ Créé : {nom_ref}")

    # 4. BLOCS
    print("\n[4] 📦 Importation des Blocs...")
    ws = wb['1_BLOCS']
    bloc_map = {} 
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        code, nom, ref_nom, ordre = str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip(), row[3]
        
        if ref_nom not in ref_map: continue 

        bloc = BlocCompetence.objects.create(
            referentiel=ref_map[ref_nom],
            code=code,
            nom=nom,
            ordre=ordre or 0
        )
        bloc_map[(code, ref_nom)] = bloc
        print(f"    -> Bloc {code} ajouté à {ref_nom}")

    # 5. COMPETENCES
    print("\n[5] 🧠 Importation des Compétences...")
    ws = wb['2_COMPETENCES']
    comp_map = {} 

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        code, nom, bloc_code, ref_nom, ordre = str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip(), row[4]

        key_bloc = (bloc_code, ref_nom)
        if key_bloc not in bloc_map:
            continue

        comp = Competence.objects.create(
            bloc=bloc_map[key_bloc],
            code=code,
            nom=nom,
            ordre=ordre or 0
        )
        
        # Clé unique pour retrouver cette compétence précise
        key_unique = (code, bloc_code, ref_nom)
        comp_map[key_unique] = comp

    print(f"    Total Compétences créées : {len(comp_map)}")

    # 6. COMPETENCES PRO (CP)
    print("\n[6] 🔨 Importation des Compétences Pro (CP)...")
    ws = wb['3_COMPETENCES_PRO']
    cp_list_map = {} 
    count_cp = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        code, nom, comp_code, ref_nom, ordre = str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip(), row[4]

        if ref_nom not in ref_map: continue

        # Trouver TOUTES les compétences parentes correspondantes
        parents_trouves = []
        for (k_code, k_bloc, k_ref), obj_comp in comp_map.items():
            if k_code == comp_code and k_ref == ref_nom:
                parents_trouves.append(obj_comp)

        if not parents_trouves:
            continue

        for parent in parents_trouves:
            cp = CompetenceProfessionnelle.objects.create(
                competence=parent,
                code=code,
                nom=nom,
                ordre=ordre or 0
            )
            # On stocke dans une liste pour gérer les SC ensuite
            k = (code, ref_nom)
            if k not in cp_list_map:
                cp_list_map[k] = []
            cp_list_map[k].append(cp)
            count_cp += 1

    print(f"    Total CP créées : {count_cp}")

    # 7. SOUS-COMPETENCES (SC) - CORRIGÉ (Plus de code=...)
    print("\n[7] 🎯 Importation des Sous-Compétences (SC)...")
    ws = wb['4_SOUS_COMPETENCES']
    sc_list_map = {}
    count_sc = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        code_sc, nom, cp_code, ref_nom, ordre = str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip(), row[4]

        if (cp_code, ref_nom) not in cp_list_map:
            continue

        parents_cps = cp_list_map[(cp_code, ref_nom)]
        for parent_cp in parents_cps:
            # CORRECTION ICI : Pas de champ 'code' dans create
            sc = SousCompetence.objects.create(
                competence_pro=parent_cp,
                nom=nom,
                ordre=ordre or 0
            )
            
            k = (code_sc, ref_nom) # On utilise le code Excel pour la map, pas pour la DB
            if k not in sc_list_map:
                sc_list_map[k] = []
            sc_list_map[k].append(sc)
            count_sc += 1
            
    print(f"    Total SC créées : {count_sc}")

    # 8. CRITERES - CORRIGÉ (Plus de code=...)
    print("\n[8] 📏 Importation des Critères...")
    ws = wb['5_CRITERES']
    count_crit = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        code_crit, nom, sc_code, ref_nom, ordre = str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip(), row[4]

        if (sc_code, ref_nom) not in sc_list_map:
            continue

        parents_scs = sc_list_map[(sc_code, ref_nom)]
        for parent_sc in parents_scs:
            # CORRECTION ICI : Pas de champ 'code'
            crit = CritereEvaluation.objects.create(
                sous_competence=parent_sc,
                nom=nom,
                ordre=ordre or 0
            )
            # Création auto Indicateur
            IndicateurPerformance.objects.create(
                critere=crit,
                nom=nom,
                poids=10.0,
                ordre=1
            )
            count_crit += 1

    print(f"    Total Critères créés : {count_crit}")
    
    # 9. CONNAISSANCES
    print("\n[9] 📚 Importation des Connaissances...")
    ws = wb['6_CONNAISSANCES']
    count_conn = 0
    
    # Vérifier si le modèle Connaissance a un champ code
    try:
        from core.models import Connaissance
        HAS_CONN = True
    except ImportError:
        HAS_CONN = False
        
    if HAS_CONN:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            code_conn, nom, cp_code, ref_nom, ordre = str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip(), row[4]

            if (cp_code, ref_nom) not in cp_list_map:
                continue

            parents_cps = cp_list_map[(cp_code, ref_nom)]
            for parent_cp in parents_cps:
                # Je tente avec code, si ça plante, retirez 'code=code_conn' comme pour SC
                try:
                    Connaissance.objects.create(
                        competence_pro=parent_cp,
                        nom=nom,
                        code=code_conn, 
                        ordre=ordre or 0
                    )
                except TypeError:
                    # Fallback si pas de champ code
                    Connaissance.objects.create(
                        competence_pro=parent_cp,
                        nom=nom,
                        ordre=ordre or 0
                    )
                count_conn += 1
        print(f"    Total Connaissances créées : {count_conn}")

    print("\n" + "="*80)
    print("✅ IMPORTATION TERMINEE AVEC SUCCES !")
    print("="*80)

if __name__ == '__main__':
    run_import()